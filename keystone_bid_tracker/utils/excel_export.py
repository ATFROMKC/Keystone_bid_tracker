"""
Keystone Bid Tracker - Excel Export
Exports bids to .xlsx matching the original backlog format.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


HEADER_FILL = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="f0f0f0")
CELL_FONT = Font(name="Segoe UI", size=11, color="333333")
BORDER = Border(
    bottom=Side(style="thin", color="dddddd"),
)


def export_bids(bids: list, filepath: str):
    """Export a list of bid dicts (from get_all_bids_for_export) to xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bids"

    headers = [
        "BID DATE", "Estimator", "BID NAME", "BID TOTAL $",
        "SOLID SURF. SF", "STONE SF", "BID TO", "Status",
        "Won By", "Rev #", "Notes"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, b in enumerate(bids, 2):
        try:
            d = datetime.strptime(b["original_bid_date"], "%Y-%m-%d")
            date_str = d.strftime("%m/%d/%Y")
        except (ValueError, TypeError):
            date_str = b.get("original_bid_date", "")

        ws.cell(row=row_idx, column=1, value=date_str).font = CELL_FONT
        ws.cell(row=row_idx, column=2, value=b.get("estimator", "")).font = CELL_FONT
        ws.cell(row=row_idx, column=3, value=b.get("bid_name", "")).font = CELL_FONT

        total_cell = ws.cell(row=row_idx, column=4, value=b.get("bid_total", 0))
        total_cell.number_format = '$#,##0.00'
        total_cell.font = CELL_FONT

        ws.cell(row=row_idx, column=5, value=b.get("solid_surf_sf", 0)).font = CELL_FONT
        ws.cell(row=row_idx, column=6, value=b.get("stone_sf", 0)).font = CELL_FONT
        ws.cell(row=row_idx, column=7, value=b.get("customer_names", "")).font = CELL_FONT
        ws.cell(row=row_idx, column=8, value=b.get("status", "")).font = CELL_FONT
        ws.cell(row=row_idx, column=9, value=b.get("won_customer_name", "") or "").font = CELL_FONT
        ws.cell(row=row_idx, column=10, value=b.get("revision_no", 1)).font = CELL_FONT
        ws.cell(row=row_idx, column=11, value=b.get("notes", "") or "").font = CELL_FONT

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = BORDER

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row in range(2, len(bids) + 2):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 4, 50)

    wb.save(filepath)


def export_report_data(report_data: dict, filepath: str):
    """Export report summary tables to xlsx."""
    wb = Workbook()

    # Status summary sheet
    ws1 = wb.active
    ws1.title = "By Status"
    ws1.append(["Status", "Count", "Total Value"])
    for item in report_data.get("by_status", []):
        ws1.append([item["status"], item["cnt"], item["total_value"]])

    # By customer sheet
    ws2 = wb.create_sheet("By Account")
    ws2.append(["Account", "Bids", "Won", "Total Value"])
    for item in report_data.get("by_customer", []):
        ws2.append([item["customer_name"], item["bid_count"],
                     item["won_count"], item["total_value"]])

    # Win rate sheet
    ws3 = wb.create_sheet("Win Rate")
    ws3.append(["Estimator", "Total Bids", "Won", "Win Rate %"])
    win_data = report_data.get("win_rate", {})
    for item in win_data.get("by_estimator", []):
        ws3.append([item["estimator"], item["total"], item["won"], item["rate"]])
    ws3.append([])
    ws3.append(["Overall Win Rate", f"{win_data.get('overall_rate', 0)}%"])

    wb.save(filepath)


def export_pm_jobs(rows: list, filepath: str):
    """Export PM Job Manager rows to xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PM Jobs"

    headers = [
        "DATE WON",
        "JOB NAME",
        "ACCOUNT",
        "SALESPERSON",
        "PM",
        "JOB TYPE",
        "BID TOTAL $",
        "INVOICE STATUS",
        "EST COMPLETE",
        "MORAWARE DATE",
        "NOTEBOOK STATUS",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        def _fmt_date(val):
            if not val:
                return ""
            try:
                return datetime.strptime(val, "%Y-%m-%d").strftime("%m/%d/%Y")
            except (ValueError, TypeError):
                return val

        ws.cell(row=row_idx, column=1, value=_fmt_date(row.get("date_won"))).font = CELL_FONT
        ws.cell(row=row_idx, column=2, value=row.get("job_name", "")).font = CELL_FONT
        ws.cell(row=row_idx, column=3, value=row.get("account", "")).font = CELL_FONT
        ws.cell(row=row_idx, column=4, value=row.get("salesperson", "")).font = CELL_FONT
        ws.cell(row=row_idx, column=5, value=row.get("project_manager", "")).font = CELL_FONT
        ws.cell(row=row_idx, column=6, value=row.get("job_type", "")).font = CELL_FONT

        total_cell = ws.cell(row=row_idx, column=7, value=row.get("bid_total", 0))
        total_cell.number_format = "$#,##0.00"
        total_cell.font = CELL_FONT

        ws.cell(row=row_idx, column=8, value=row.get("invoice_status", "")).font = CELL_FONT
        ws.cell(row=row_idx, column=9, value=_fmt_date(row.get("est_complete_date"))).font = CELL_FONT
        ws.cell(row=row_idx, column=10, value=_fmt_date(row.get("moraware_date"))).font = CELL_FONT
        ws.cell(row=row_idx, column=11, value=row.get("notebook_status", "")).font = CELL_FONT

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = BORDER

    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row in range(2, len(rows) + 2):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 4, 50)

    wb.save(filepath)
