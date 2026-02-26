"""
Keystone Bid Tracker - Excel Import
Import historical backlog from Bid_Tracker_Backlog.xlsx.
"""

from openpyxl import load_workbook
from datetime import datetime


# Expected column headers (case-insensitive matching)
COLUMN_MAP = {
    "bid date": "bid_date",
    "estimator": "estimator",
    "bid name": "bid_name",
    "bid total $": "bid_total",
    "bid total": "bid_total",
    "solid surf. sf": "solid_surf_sf",
    "solid surf sf": "solid_surf_sf",
    "stone sf": "stone_sf",
    "bid to": "bid_to_1",
    "bid to2": "bid_to_2",
    "bid to3": "bid_to_3",
    "bid to4": "bid_to_4",
    "bid won?": "bid_won",
}


def _normalize_header(val):
    if val is None:
        return ""
    return str(val).strip().lower()


def _parse_date(val):
    """Try to parse a date from various formats, return ISO string or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%m/%d/%y", "%m-%d-%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _parse_float(val):
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return 0.0


def preview_import(filepath: str) -> list:
    """Read Excel file and return list of parsed row dicts for preview."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(min_row=1)
    header_row = next(rows_iter)

    col_indices = {}
    for idx, cell in enumerate(header_row):
        norm = _normalize_header(cell.value)
        if norm in COLUMN_MAP:
            col_indices[COLUMN_MAP[norm]] = idx

    parsed = []
    for row in rows_iter:
        vals = [cell.value for cell in row]

        bid_name = str(vals[col_indices.get("bid_name", 0)] or "").strip() if "bid_name" in col_indices else ""
        if not bid_name:
            continue

        bid_date = _parse_date(vals[col_indices["bid_date"]]) if "bid_date" in col_indices else None
        estimator = str(vals[col_indices.get("estimator", 0)] or "").strip() if "estimator" in col_indices else ""
        bid_total = _parse_float(vals[col_indices["bid_total"]]) if "bid_total" in col_indices else 0
        solid_sf = _parse_float(vals[col_indices["solid_surf_sf"]]) if "solid_surf_sf" in col_indices else 0
        stone_sf = _parse_float(vals[col_indices["stone_sf"]]) if "stone_sf" in col_indices else 0

        customers = []
        for key in ("bid_to_1", "bid_to_2", "bid_to_3", "bid_to_4"):
            if key in col_indices:
                c = str(vals[col_indices[key]] or "").strip()
                if c:
                    customers.append(c)

        bid_won_raw = str(vals[col_indices.get("bid_won", 0)] or "").strip().lower() if "bid_won" in col_indices else ""
        won = bid_won_raw in ("yes", "y", "1", "true")

        parsed.append({
            "bid_name": bid_name,
            "bid_date": bid_date or "",
            "estimator": estimator,
            "bid_total": bid_total,
            "solid_surf_sf": solid_sf,
            "stone_sf": stone_sf,
            "customers": customers,
            "won": won,
        })

    wb.close()
    return parsed


def check_duplicates(db, parsed_rows: list) -> list:
    """Return a bool flag list where True means the row already exists."""
    duplicate_flags = []
    for row in parsed_rows:
        bid_name = row["bid_name"]
        bid_date = row["bid_date"]
        estimator = row["estimator"]
        is_duplicate = bool(bid_date) and db.bid_exists(bid_name, bid_date, estimator)
        duplicate_flags.append(is_duplicate)
    return duplicate_flags


def analyze_rows(db, parsed_rows: list) -> list:
    """
    Analyze parsed rows for preview status.
    Returns list of dicts with duplicate/date/customer flags and reason labels.
    """
    analysis = []
    for row in parsed_rows:
        bid_name = row["bid_name"]
        bid_date = row["bid_date"]
        estimator = row["estimator"]
        has_valid_date = bool(bid_date)
        has_customers = len(row["customers"]) > 0
        is_duplicate = has_valid_date and db.bid_exists(bid_name, bid_date, estimator)

        reasons = []
        if is_duplicate:
            reasons.append("Duplicate")
        if not has_valid_date:
            reasons.append("Invalid Date")
        if not has_customers:
            reasons.append("No Customer")

        analysis.append({
            "is_duplicate": is_duplicate,
            "has_valid_date": has_valid_date,
            "has_customers": has_customers,
            "reasons": reasons,
        })

    return analysis


def run_import(db, parsed_rows: list, progress_callback=None, import_duplicates: bool = False) -> dict:
    """
    Import parsed rows into the database.
    Returns summary dict: {imported, customers_created, skipped}.
    """
    imported = 0
    customers_created = 0
    skipped = 0
    total = len(parsed_rows)

    for i, row in enumerate(parsed_rows):
        if progress_callback:
            progress_callback(i + 1, total)

        bid_name = row["bid_name"]
        bid_date = row["bid_date"]
        estimator = row["estimator"]

        if not bid_date:
            skipped += 1
            continue

        if not import_duplicates and db.bid_exists(bid_name, bid_date, estimator):
            skipped += 1
            continue

        # Resolve customers
        customer_ids = []
        for cname in row["customers"]:
            existing = db.get_customer_by_name(cname)
            if existing:
                customer_ids.append(existing["id"])
            else:
                cid = db.add_customer(cname)
                customer_ids.append(cid)
                customers_created += 1

        if not customer_ids:
            skipped += 1
            continue

        bid_id = db.add_bid(
            bid_name=bid_name,
            estimator=estimator,
            original_bid_date=bid_date,
            notes="",
            customer_ids=customer_ids,
            bid_total=row["bid_total"],
            solid_surf_sf=row["solid_surf_sf"],
            stone_sf=row["stone_sf"],
        )

        if row["won"] and customer_ids:
            db.mark_bid_status(bid_id, "WON", customer_ids[0])

        imported += 1

    return {
        "imported": imported,
        "customers_created": customers_created,
        "skipped": skipped,
    }
